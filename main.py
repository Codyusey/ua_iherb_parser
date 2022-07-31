"""
One-thread parser  'https://ua.iherb.com/'
Parse only from ukrainian ip
Parsing speed - 3 items per minute
OS: Windows 10
"""
import sys
import undetected_chromedriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from bs4 import BeautifulSoup
import json
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
import datetime
import time
import random
import os
import glob
import os.path
from tqdm import tqdm
from winsound import MessageBeep, MB_OK, MB_ICONHAND
import colorama
from colorama import Fore
import cursor

colorama.init(autoreset=True)
current_data_time = datetime.datetime.now().strftime("%d_%m_%Y_%H_%M")
current_data = datetime.datetime.now().strftime("%d_%m_%Y")
abs_path_work_dir = os.path.dirname((os.path.abspath(__file__)))
input_dir = abs_path_work_dir + "\\Input\\"
output_dir = abs_path_work_dir + "\\Output\\"
temp_dir = abs_path_work_dir + "\\Temp\\"
filename_data_json = temp_dir + "loaded_codes.json"
URL = "https://ua.iherb.com/"
header_items = ["1. Код з сайту", "2. Цiна, $", "3. Заощадьте у кошику, %", "4. Акційна ціна, $",
                "5. Економія, ціна, $", "6. Економія, %", "7. Код УТК", "8. Наявність", "9. Термін придатності"]
full_list_codes = []
items_dict = {}
error_codes = []
message_len = 80
driver = None


def get_name_xlsx(file_extension='xlsx') -> str:
    filelist = glob.glob(os.path.join(input_dir, f"*.{file_extension}"))
    if len(filelist) == 0:
        print('Щоб розпочати парсинг сайту https://ua.iherb.com/ '
              '\nдодайте один файл даних з розширенням "xlsx" до папки "Input"')
        sys.exit()
    elif 3 > len(filelist) > 1:
        for name in filelist:
            if '~$' not in name:
                input_filename_xlsx_ = name
                return input_filename_xlsx_
        print('Щоб розпочати парсинг сайту https://ua.iherb.com/ '
              '\nу папці "Input" має бути лише один файл даних з розширенням "xlsx"')
        sys.exit()
    else:
        input_filename_xlsx_ = filelist[0]
        return input_filename_xlsx_


def print_ln(message, tab_type='…', start_ln='\n', end_ln='\n', color=Fore.WHITE):
    tab_len = int((message_len - 2 - len(message)) / 2)
    print(color + start_ln + tab_type * tab_len + " " + message + " " + ' ' * (len(message) % 2) + tab_type * tab_len,
          end=end_ln)


def beep(times=1, b_type=MB_OK):
    for i in range(times):
        MessageBeep(b_type)
        time.sleep(1)


def timer_countdown(num_of_secs):
    beep()
    print()
    cursor.hide()
    while num_of_secs + 1:
        m, s = divmod(num_of_secs, 60)
        min_sec_format = '{:02d}:{:02d} minutes.'.format(m, s)
        print(Fore.LIGHTGREEN_EX + f"\rRestart parser to pass captcha in {min_sec_format}", end="")
        time.sleep(1)
        num_of_secs -= 1


def data_from_xlsx(file_name_xlsx: str) -> list:
    codes_li = []
    wb = load_workbook(file_name_xlsx)
    sheet_ranges = wb.active
    column_a = sheet_ranges['A']
    for cell in column_a:
        if cell.value is not None:
            codes_li.append(cell.value)
    if codes_li == [None]:
        print_ln(f'Помилка читання списку кодів з файлу "{get_name_xlsx()}". '
                 f'\nКоди мають бути записані в перший стовпець.', start_ln='', color=Fore.RED)
        sys.exit()
    return codes_li


def write_data_json(filename, data):
    try:
        with open(filename, 'w', encoding="utf-8") as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception:
        print(f'Error write - {filename}')


def load_data_json(filename=filename_data_json):
    try:
        with open(filename, 'r', encoding="utf-8") as f:
            data_json = json.load(f)
            return data_json
    except Exception:
        print(f'Error read "{filename}"')
        write_data_json(filename, {})  # clear cache
        return {}


def parser():
    global driver
    while len(items_dict) != len(full_list_codes):
        repeat = False
        try:
            driver = undetected_chromedriver.Chrome()
            time.sleep(2)
            driver.maximize_window()
            driver.get(URL)
            time.sleep(random.randrange(5, 6))
        except Exception:
            print_ln('Webdriver initialization error. Trying again...', start_ln='\r', end_ln='\r', tab_type='*',
                     color=Fore.RED)
            beep(b_type=MB_ICONHAND)
            time.sleep(3)
            try:
                driver.quit()
                time.sleep(3)
            except Exception:
                sys.exit()
            continue
        try:
            driver.maximize_window()
            driver.find_element(
                By.XPATH, '/html/body/header/div[1]/div[1]/div[2]/div/div[2]/div/div[2]/button[1]').click()
            time.sleep(random.randrange(5, 6))
        except Exception:
            pass
        finally:
            for code in tqdm(full_list_codes, desc='Scraping pages', unit='page', ncols=message_len):
                if code in items_dict:
                    continue
                while True:
                    items_lists = [''] * len(header_items)
                    real_code = ""
                    try:
                        driver.maximize_window()
                        driver.find_element(By.ID, "px-captcha")
                        repeat = True
                        break
                    except Exception:
                        pass
                    try:
                        driver.maximize_window()
                        search_input = driver.find_element(
                            By.XPATH, "/html/body/header/div[1]/div[3]/div[2]/div[1]/div[2]/div[1]/form/div/input")
                        search_input.send_keys(code + Keys.ENTER)
                        time.sleep(random.randrange(5, 6))
                        driver.maximize_window()
                        soup = BeautifulSoup(driver.page_source, "html.parser")  # "lxml"
                        try:
                            real_code = soup.find("ul", {"id": "product-specs-list"}).find("span", {
                                "itemprop": "sku"}).text.strip()
                            if '-' not in code:
                                real_code = real_code.replace('-', "")
                        except Exception as ex:
                            print(ex)

                        if real_code == code:
                            items_lists[0] = real_code
                            try:
                                try:
                                    soup.find("div", {"class": "discount-in-cart"}).text.strip()
                                    price = soup.find("div", {"id": "price"}).text.strip()
                                    price = price.split('\n')[-1].strip()
                                    items_lists[1] = float(price[1:])
                                except Exception:
                                    price = soup.find("div", {"id": "price"}).text.strip()
                                    items_lists[1] = float(price[1:])
                            except Exception:
                                pass

                            try:
                                discount_in_cart = soup.find("div", {"class": "discount-in-cart"}).text.strip()
                                items_lists[2] = int(discount_in_cart.split()[1]) / 100
                            except Exception:
                                pass
                            try:
                                disc_price = soup.find("b", {"class": "s24"}).text.strip()
                                items_lists[3] = float(disc_price[1:])
                            except Exception:
                                pass
                            try:
                                econom_price = soup.find("section", {"id": "product-discount"}).text.strip()
                                econom_price = econom_price.split()[1]
                                items_lists[4] = float(econom_price[1:])
                            except Exception:
                                pass
                            try:
                                econom_price2 = soup.find("span", {"class": "discount-text"}).text.strip()
                                items_lists[5] = int(econom_price2) / 100
                            except Exception:
                                pass
                            try:
                                bar_code_utk = soup.find("ul", {"id": "product-specs-list"}
                                                         ).find("span", {"itemprop": "gtin12"}).text.strip()
                                items_lists[6] = bar_code_utk
                            except Exception:
                                pass
                            try:
                                in_stock = soup.find("div", {"id": "stock-status"}).text.strip()
                                in_stock = in_stock.split('\n')[0]
                                items_lists[7] = in_stock
                            except Exception:
                                pass
                            try:
                                guarantee = soup.find("ul", {"id": "product-specs-list"}).find_next().text.strip()
                                guarantee = 'до ' + ((guarantee.split('\n')[0]).split('?')[1]).strip()
                                items_lists[8] = guarantee
                            except Exception:
                                pass
                            finally:
                                items_dict[code] = items_lists
                            try:
                                write_data_json(filename_data_json, items_dict)
                                break
                            except Exception:
                                print(f'Error write "{filename_data_json}"')
                        else:
                            print_ln("Увага! Парсер працює, якщо вікно браузера завжди поверх усіх вікон.",
                                     tab_type='*', start_ln='\r', end_ln='\r', color=Fore.RED)
                            beep(b_type=MB_ICONHAND)
                            time.sleep(1)
                            continue
                    except Exception as ex:
                        print(ex)
                        pass
                    try:
                        driver.maximize_window()
                        notfound_type1 = driver.find_element(
                            By.XPATH, "/html/body/div[6]/div[1]/div[4]/div[1]/div/div/p").text
                        print_ln(f'Error parsing item with code"{code}"', tab_type='-', start_ln='\r', end_ln='\r',
                                 color=Fore.RED)
                        beep(b_type=MB_ICONHAND)
                        items_lists[0] = f'Не вдалося знайти жодного товару, що відповідає запиту: "{code}"'
                        items_dict[code] = items_lists
                        write_data_json(filename_data_json, items_dict)
                        driver.get(URL)
                        time.sleep(random.randrange(4, 5))
                        break
                    except Exception:
                        pass
                    try:
                        driver.maximize_window()
                        notfound_type2 = driver.find_element(By.XPATH, "/html/body/div[6]/div/div[2]/h1").text
                        print_ln(f'Error parsing item with code"{code}"', tab_type='-', start_ln='\r', end_ln='\r',
                                 color=Fore.RED)
                        beep(b_type=MB_ICONHAND)
                        items_lists[0] = f'Не вдалося знайти жодного товару, що відповідає запиту: "{code}"'
                        items_dict[code] = items_lists
                        write_data_json(filename_data_json, items_dict)
                        driver.get(URL)
                        time.sleep(random.randrange(4, 5))
                        break
                    except Exception:
                        pass

                if repeat:
                    driver.quit()
                    waiting_time = 11
                    timer_countdown(waiting_time * 60)
                    os.system("cls")
                    print_ln("iHerb_ua parser's continues scraping")
                    cursor.hide()
                    driver.quit()
                    break


def write_items_xlsx(name_xlsx, data_dict_, header_list_):
    wb = Workbook()
    ws = wb.active
    ws.column_dimensions['A'].width = len(header_list_[0])
    ws.column_dimensions['B'].width = len(header_list_[1])
    ws.column_dimensions['C'].width = len(header_list_[2])
    ws.column_dimensions['D'].width = len(header_list_[3])
    ws.column_dimensions['E'].width = len(header_list_[4]) - 1
    ws.column_dimensions['F'].width = len(header_list_[5]) + 1
    ws.column_dimensions['G'].width = len(header_list_[6]) + 3
    ws.column_dimensions['H'].width = len(header_list_[7]) + 2
    ws.column_dimensions['I'].width = len(header_list_[8]) + 2
    ws.title = f'ua_iherb_{len(full_list_codes)}codes_{current_data}'
    if header_list_ != "":
        ws.append(header_list_)
    if len(data_dict_) != 0:
        for row in data_dict_.values():
            ws.append(row)
        for i in ('A', 'C', 'F', 'G', 'H', 'I'):
            col_range = ws[i]
            for cell in col_range:
                cell.font = Font(name='Calibri', size=11, bold=False)
                cell.alignment = Alignment(horizontal='general', vertical="center")
                cell.number_format = '0%'
        for cell in ws["1:1"]:
            cell.font = Font(name='Calibri', size=11, bold=True)

        wb.save(name_xlsx)
    else:
        print(f'Noting to write. Data is empty')


def print_error_codes():
    for code, val in items_dict.items():
        if val[0] == f'Не вдалося знайти жодного товару, що відповідає запиту: "{code}"':
            error_codes.append(code)
    if len(error_codes) != 0:
        print_ln(f"Error parsing codes:", tab_type='', start_ln='', color=Fore.RED)
        for code in error_codes:
            print_ln(code, tab_type='', start_ln='', color=Fore.RED)


if __name__ == '__main__':
    start_time = time.time()
    cursor.hide()
    try:
        full_list_codes = data_from_xlsx(get_name_xlsx())
    except Exception:
        print(f'Error open data from file : {get_name_xlsx()}')
    items_dict = load_data_json()

    if not set(full_list_codes) >= set(items_dict):
        write_data_json(filename_data_json, {})  # clear cache
        items_dict = load_data_json()

    if len(items_dict) == 0:
        print_ln("iHerb_ua parser's started")
    else:
        print_ln("iHerb_ua parser's continues scraping")
    parser()
    xlsx_pathname = output_dir + 'ua_iherb_' + str(len(full_list_codes)) + 'codes_' + current_data_time + '.xlsx'
    write_items_xlsx(xlsx_pathname, load_data_json(), header_items)
    write_data_json(filename_data_json, {})  # clear cache
    print_error_codes()
    print_ln('End', start_ln='')
    os.startfile(output_dir)
    beep(3)
    finish_time = (time.time() - start_time)
    print(f"Time spent on the scraping : {int(finish_time / 60)}m:{int(finish_time % 60)}s")
    cursor.show()
